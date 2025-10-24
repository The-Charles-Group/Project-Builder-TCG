
#!/usr/bin/env python3
"""
Convert MASTER_CONTROL_ROOM.md to Excel format with proper sections and tables
"""

import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def parse_markdown_to_excel(md_file: str, output_file: str):
    """Parse markdown file and create structured Excel workbook"""
    
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Split by main sections (## headers)
    sections = re.split(r'\n## ', content)
    
    # Process each section
    for section in sections[1:]:  # Skip first empty split
        lines = section.split('\n')
        section_title = lines[0].strip()
        section_content = '\n'.join(lines[1:])
        
        # Create safe sheet name (Excel limit: 31 chars, no special chars)
        sheet_name = section_title[:31].replace('/', '-').replace('\\', '-').replace(':', '-')
        ws = wb.create_sheet(title=sheet_name)
        
        # Add section title as header
        ws['A1'] = section_title
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Parse subsections and content
        current_row = 3
        subsections = re.split(r'\n### ', section_content)
        
        for subsection in subsections:
            if not subsection.strip():
                continue
                
            sub_lines = subsection.split('\n')
            sub_title = sub_lines[0].strip()
            sub_content = '\n'.join(sub_lines[1:])
            
            # Add subsection header
            ws[f'A{current_row}'] = sub_title
            ws[f'A{current_row}'].font = Font(size=12, bold=True, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            current_row += 1
            
            # Check for tables (markdown tables with | separators)
            table_pattern = r'\|(.+)\|'
            table_matches = re.findall(table_pattern, sub_content, re.MULTILINE)
            
            if table_matches and len(table_matches) > 1:
                # Parse markdown table
                headers = [h.strip() for h in table_matches[0].split('|') if h.strip()]
                
                # Write headers
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                current_row += 1
                
                # Skip separator row (---|----|----)
                data_rows = [r for r in table_matches[2:] if not re.match(r'^[-:\s|]+$', r)]
                
                # Write data rows
                for row_data in data_rows:
                    values = [v.strip() for v in row_data.split('|') if v.strip()]
                    for col_idx, value in enumerate(values, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    current_row += 1
                
                current_row += 1  # Add spacing after table
                
            elif '```' in sub_content:
                # Handle code blocks
                code_blocks = re.findall(r'```(.+?)```', sub_content, re.DOTALL)
                for code_block in code_blocks:
                    # Write code block header
                    ws[f'A{current_row}'] = 'Code Block:'
                    ws[f'A{current_row}'].font = Font(italic=True)
                    current_row += 1
                    
                    # Write code content
                    ws[f'A{current_row}'] = code_block.strip()
                    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    ws[f'A{current_row}'].font = Font(name='Consolas', size=9)
                    ws[f'A{current_row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    ws.row_dimensions[current_row].height = max(15, len(code_block.split('\n')) * 15)
                    current_row += 2
            else:
                # Regular text content
                lines = sub_content.strip().split('\n')
                for line in lines:
                    line = line.strip()
                    if line and not line.startswith('│') and not line.startswith('┌'):
                        ws[f'A{current_row}'] = line
                        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                        current_row += 1
                
                current_row += 1  # Add spacing
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Excel file created: {output_file}")
    print(f"📊 Total sheets: {len(wb.sheetnames)}")
    print(f"📄 Sheets: {', '.join(wb.sheetnames[:5])}{'...' if len(wb.sheetnames) > 5 else ''}")

if __name__ == "__main__":
    parse_markdown_to_excel(
        md_file="MASTER_CONTROL_ROOM.md",
        output_file="MASTER_CONTROL_ROOM.xlsx"
    )
